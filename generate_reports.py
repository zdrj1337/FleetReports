"""
Generate weekly driver payment reports from the normalised fleet data.

For each driver it writes an Excel report that shows the per-platform breakdown
(Uber / Bolt) and a summary box, and for the whole fleet it writes one summary
file with a bar chart and a "By Platform" breakdown sheet.
Derived values (firm commission, net earnings, bank transfer, totals) are
written as real Excel FORMULAS so the reports recalculate if the numbers change
and the math is transparent / auditable.

Usage
-----
    python generate_reports.py 2026-W10        # single period
    python generate_reports.py --all           # every period found under data/
"""

import os
import re
import sys
from datetime import datetime

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

import fleet_logic

# ── Business rules ────────────────────────────────────────────────────────────
FIRM_RATE  = fleet_logic.FIRM_RATE
FLEET_NAME = "MASTER FLEET"

# ── Palette ───────────────────────────────────────────────────────────────────
FONT   = "Arial"
NAVY   = "1F4E79"
LIGHT  = "D6E4F0"
INFO   = "F0F5FF"
ALT    = "EEF3FF"

# Thick left-border accent colour per platform
PLATFORM_ACCENT = {"Bolt": "1A7A3E", "Uber": "1C1C1C"}

# Currency: positives normal, negatives in parentheses, zeros as dash
CUR = '#,##0.00" RON";(#,##0.00)" RON";"-"'
PCT = "0%"

# ── Reusable style objects ────────────────────────────────────────────────────
title_font  = Font(name=FONT, bold=True, size=14, color="FFFFFF")
hdr_font    = Font(name=FONT, bold=True, color="FFFFFF")
bold        = Font(name=FONT, bold=True)
normal      = Font(name=FONT)
subtitle    = Font(name=FONT, italic=True, color="595959")
navy_fill   = PatternFill("solid", fgColor=NAVY)
light_fill  = PatternFill("solid", fgColor=LIGHT)
info_fill   = PatternFill("solid", fgColor=INFO)
alt_fill    = PatternFill("solid", fgColor=ALT)
center      = Alignment(horizontal="center", vertical="center")
thin        = Side(style="thin", color="BFBFBF")
border      = Border(left=thin, right=thin, top=thin, bottom=thin)

HEADERS = [
    "Platform", "Gross", "Platform commission", "Cash collected",
    "Tips", "Bonus", "Firm commission (7%)", "Bank transfer",
]


# ── Helpers ───────────────────────────────────────────────────────────────────

def _header_row(ws, row: int, headers: list[str]) -> None:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font      = hdr_font
        c.fill      = navy_fill
        c.alignment = center
        c.border    = border


def _fit_one_page(ws) -> None:
    """Print / export onto a single landscape page (unlimited height)."""
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)


def _outline_border(ws, min_row: int, max_row: int,
                    min_col: int, max_col: int,
                    style: str = "thin", color: str = "BFBFBF") -> None:
    """Draw the outer rectangular border around a range, preserving inner borders."""
    s = Side(style=style, color=color)
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            existing = ws.cell(row=r, column=c).border
            ws.cell(row=r, column=c).border = Border(
                top    = s           if r == min_row else existing.top,
                bottom = s           if r == max_row else existing.bottom,
                left   = s           if c == min_col else existing.left,
                right  = s           if c == max_col else existing.right,
            )


def _platform_left(platform: str) -> Side:
    color = PLATFORM_ACCENT.get(platform, "000000")
    return Side(style="medium", color=color)


# ── Driver report ─────────────────────────────────────────────────────────────

def write_driver_report(driver: str, df, period: str, out_path: str) -> None:
    rows = df[df["driver"] == driver].reset_index(drop=True)
    n    = len(rows)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    # Row 1–2: title banner
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = FLEET_NAME;  c.font = title_font; c.fill = navy_fill; c.alignment = center

    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = "Weekly Driver Payment Report"
    c.font = subtitle; c.alignment = center

    # Rows 4–7: info section
    info_rows = [
        ("Driver:",               driver),
        ("Period:",               period),
        ("Firm commission rate:", FIRM_RATE),
        ("Generated:",            datetime.now().strftime("%d.%m.%Y  %H:%M")),
    ]
    for i, (label, value) in enumerate(info_rows):
        r = 4 + i
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = bold; lc.fill = info_fill
        vc = ws.cell(row=r, column=2, value=value)
        vc.font = normal; vc.fill = info_fill
        if label == "Firm commission rate:":
            vc.number_format = PCT
    _outline_border(ws, 4, 7, 1, 2)

    # Row 8: table header + freeze
    hrow = 8
    _header_row(ws, hrow, HEADERS)
    ws.freeze_panes = f"A{hrow + 1}"

    # Data rows
    first = hrow + 1
    for j, r in rows.iterrows():
        rr        = first + j
        row_fill  = alt_fill if j % 2 == 1 else None
        plat      = r["platform"]

        c1 = ws.cell(row=rr, column=1, value=plat)
        c1.font   = normal
        c1.border = Border(left=_platform_left(plat), right=thin, top=thin, bottom=thin)
        if row_fill:
            c1.fill = row_fill

        values = [
            float(r["gross"]),
            float(r["platform_commission"]),
            float(r["cash"]),
            float(r["tips"]),
            float(r["bonus"]),
            f"=B{rr}*$B$6",                              # firm commission
            f"=B{rr}-C{rr}-D{rr}+E{rr}+F{rr}-G{rr}",  # bank transfer
        ]
        for col, val in enumerate(values, start=2):
            c = ws.cell(row=rr, column=col, value=val)
            c.number_format = CUR
            c.font          = normal
            c.border        = border
            if row_fill:
                c.fill = row_fill

    # Total row
    last = first + n - 1
    trow = last + 1
    tc = ws.cell(row=trow, column=1, value="TOTAL")
    tc.font = bold; tc.fill = light_fill; tc.border = border
    for col in range(2, 9):
        L = get_column_letter(col)
        c = ws.cell(row=trow, column=col, value=f"=SUM({L}{first}:{L}{last})")
        c.number_format = CUR; c.font = bold; c.fill = light_fill; c.border = border

    # Payment summary box
    s = trow + 2

    ws.merge_cells(f"A{s - 1}:B{s - 1}")
    shdr = ws.cell(row=s - 1, column=1, value="PAYMENT SUMMARY")
    shdr.font = hdr_font; shdr.fill = navy_fill; shdr.alignment = center

    summary_items = [
        ("Gross earnings",           f"=B{trow}"),
        ("Platform commission",      f"=C{trow}"),
        ("Firm commission (7%)",     f"=G{trow}"),
        ("Net earnings",             f"=B{trow}-C{trow}-G{trow}"),
        ("Tips",                     f"=E{trow}"),
        ("Bonus",                    f"=F{trow}"),
        ("Cash already collected",   f"=D{trow}"),
        ("Bank transfer to driver",  f"=H{trow}"),
    ]
    HIGHLIGHT = {"Net earnings", "Bank transfer to driver"}
    for k, (label, formula) in enumerate(summary_items):
        rr  = s + k
        hl  = label in HIGHLIGHT
        f   = light_fill if hl else info_fill
        lc  = ws.cell(row=rr, column=1, value=label)
        vc  = ws.cell(row=rr, column=2, value=formula)
        lc.font = vc.font = Font(name=FONT, bold=hl)
        lc.fill = vc.fill = f
        vc.number_format = CUR

    _outline_border(ws, s, s + len(summary_items) - 1, 1, 2)

    # Column widths
    for col, w in {"A": 24, "B": 15, "C": 21, "D": 16,
                   "E": 13,  "F": 13, "G": 21, "H": 16}.items():
        ws.column_dimensions[col].width = w

    _fit_one_page(ws)
    wb.save(out_path)


# ── Platform breakdown sheet ──────────────────────────────────────────────────

def _write_platform_sheet(wb, df, period: str) -> None:
    ws = wb.create_sheet("By Platform")

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "Platform Breakdown"; c.font = title_font; c.fill = navy_fill; c.alignment = center

    ws["A2"] = "Period:";              ws["A2"].font = bold
    ws["B2"] = period;                 ws["B2"].font = normal
    ws["A3"] = "Firm commission rate:"; ws["A3"].font = bold
    ws["B3"] = FIRM_RATE;              ws["B3"].number_format = PCT; ws["B3"].font = normal

    plat_hdrs = [
        "Platform", "Active Drivers", "Gross", "Platform Commission",
        "Firm Commission (7%)", "Net Earnings",
        "Tips", "Bonus", "Cash Collected", "Bank Transfer",
    ]
    hrow = 5
    _header_row(ws, hrow, plat_hdrs)
    ws.freeze_panes = f"A{hrow + 1}"

    agg = (
        df.groupby("platform")[["gross", "platform_commission", "cash", "tips", "bonus"]]
          .sum()
          .reset_index()
    )
    n_drv = df.groupby("platform")["driver"].nunique().rename("n_drv").reset_index()
    agg   = agg.merge(n_drv, on="platform").sort_values("platform").reset_index(drop=True)

    first = hrow + 1
    for j, row in agg.iterrows():
        rr      = first + j
        rf      = alt_fill if j % 2 == 1 else None
        plat    = row["platform"]

        c1 = ws.cell(row=rr, column=1, value=plat)
        c1.font   = Font(name=FONT, bold=True, color=PLATFORM_ACCENT.get(plat, "000000"))
        c1.border = Border(left=_platform_left(plat), right=thin, top=thin, bottom=thin)
        if rf:
            c1.fill = rf

        ws.cell(row=rr, column=2).value = int(row["n_drv"])
        ws.cell(row=rr, column=3).value = float(row["gross"])
        ws.cell(row=rr, column=4).value = float(row["platform_commission"])
        ws.cell(row=rr, column=5).value = f"=C{rr}*$B$3"           # firm commission
        ws.cell(row=rr, column=6).value = f"=C{rr}-D{rr}-E{rr}"    # net
        ws.cell(row=rr, column=7).value = float(row["tips"])
        ws.cell(row=rr, column=8).value = float(row["bonus"])
        ws.cell(row=rr, column=9).value = float(row["cash"])
        ws.cell(row=rr, column=10).value = f"=F{rr}+G{rr}+H{rr}-I{rr}"  # bank transfer

        for col in range(2, 11):
            c = ws.cell(row=rr, column=col)
            c.font   = normal
            c.border = border
            c.number_format = "0" if col == 2 else CUR
            if rf:
                c.fill = rf

    last = first + len(agg) - 1
    trow = last + 1
    tc = ws.cell(row=trow, column=1, value="TOTAL")
    tc.font = bold; tc.fill = light_fill; tc.border = border
    ws.cell(row=trow, column=2, value=f"=SUM(B{first}:B{last})")
    ws.cell(row=trow, column=2).number_format = "0"
    ws.cell(row=trow, column=2).font  = bold
    ws.cell(row=trow, column=2).fill  = light_fill
    ws.cell(row=trow, column=2).border = border
    for col in range(3, 11):
        L = get_column_letter(col)
        c = ws.cell(row=trow, column=col, value=f"=SUM({L}{first}:{L}{last})")
        c.number_format = CUR; c.font = bold; c.fill = light_fill; c.border = border

    for col, w in {"A": 22, "B": 16, "C": 16, "D": 24,
                   "E": 22, "F": 15, "G": 12, "H": 12, "I": 16, "J": 15}.items():
        ws.column_dimensions[col].width = w

    _fit_one_page(ws)


# ── Fleet summary ─────────────────────────────────────────────────────────────

def write_fleet_summary(df, period: str, out_path: str) -> None:
    g = (
        df.groupby("driver", as_index=False)[["gross", "platform_commission", "cash", "tips", "bonus"]]
          .sum()
          .sort_values("driver")
          .reset_index(drop=True)
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fleet summary"

    # Row 1: title
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = f"{FLEET_NAME}  —  Fleet Summary"
    c.font = title_font; c.fill = navy_fill; c.alignment = center

    # Row 2: subtitle with meta
    ws.merge_cells("A2:I2")
    c = ws["A2"]
    c.value = (
        f"Period: {period}   |   "
        f"{len(g)} drivers   |   "
        f"Generated: {datetime.now().strftime('%d.%m.%Y  %H:%M')}"
    )
    c.font = subtitle; c.alignment = center

    # Row 3: firm rate cell (referenced by formulas below)
    ws["A3"] = "Firm commission rate:"; ws["A3"].font = bold
    ws["B3"] = FIRM_RATE; ws["B3"].number_format = PCT; ws["B3"].font = normal

    # Row 5: table header
    headers = [
        "Driver", "Gross", "Platform commission", "Firm commission (7%)",
        "Net earnings", "Tips", "Bonus", "Cash collected", "Bank transfer",
    ]
    hrow = 5
    _header_row(ws, hrow, headers)
    ws.freeze_panes = f"A{hrow + 1}"
    ws.auto_filter.ref = f"A{hrow}:{get_column_letter(9)}{hrow}"  # will expand after last row

    first = hrow + 1
    for j, row in g.iterrows():
        rr = first + j
        rf = alt_fill if j % 2 == 1 else None

        c1 = ws.cell(row=rr, column=1, value=row["driver"])
        c1.font = normal; c1.border = border
        if rf:
            c1.fill = rf

        ws.cell(row=rr, column=2).value = float(row["gross"])
        ws.cell(row=rr, column=3).value = float(row["platform_commission"])
        ws.cell(row=rr, column=4).value = f"=B{rr}*$B$3"            # firm commission
        ws.cell(row=rr, column=5).value = f"=B{rr}-C{rr}-D{rr}"     # net earnings
        ws.cell(row=rr, column=6).value = float(row["tips"])
        ws.cell(row=rr, column=7).value = float(row["bonus"])
        ws.cell(row=rr, column=8).value = float(row["cash"])
        ws.cell(row=rr, column=9).value = f"=E{rr}+F{rr}+G{rr}-H{rr}"  # bank transfer

        for col in range(2, 10):
            c = ws.cell(row=rr, column=col)
            c.number_format = CUR; c.font = normal; c.border = border
            if rf:
                c.fill = rf

    last = first + len(g) - 1

    # Expand auto-filter to actual data range
    ws.auto_filter.ref = f"A{hrow}:{get_column_letter(9)}{last}"

    # Total row
    trow = last + 1
    tc = ws.cell(row=trow, column=1, value="TOTAL")
    tc.font = bold; tc.fill = light_fill; tc.border = border
    for col in range(2, 10):
        L = get_column_letter(col)
        c = ws.cell(row=trow, column=col, value=f"=SUM({L}{first}:{L}{last})")
        c.number_format = CUR; c.font = bold; c.fill = light_fill; c.border = border

    # Column widths
    for col, w in {"A": 22, "B": 16, "C": 22, "D": 22, "E": 15,
                   "F": 12, "G": 12, "H": 16, "I": 15}.items():
        ws.column_dimensions[col].width = w

    # Bar chart — bank transfer per driver
    if len(g) > 0:
        chart = BarChart()
        chart.type     = "col"
        chart.grouping = "clustered"
        chart.title    = "Bank Transfer per Driver (RON)"
        chart.y_axis.title = "RON"
        chart.style    = 10
        chart.width    = 24
        chart.height   = 14

        data_ref = Reference(ws, min_col=9, max_col=9, min_row=hrow, max_row=last)
        cats_ref = Reference(ws, min_col=1, max_col=1, min_row=first, max_row=last)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, f"A{trow + 2}")

    _fit_one_page(ws)

    # Second sheet: platform breakdown
    _write_platform_sheet(wb, df, period)

    wb.save(out_path)


# ── CLI ───────────────────────────────────────────────────────────────────────

def _find_all_periods(base: str) -> list[str]:
    """Scan the data/ tree and return every valid period string, sorted."""
    data_root = os.path.join(base, "data")
    periods: set[str] = set()
    if not os.path.isdir(data_root):
        return []
    for entry in os.listdir(data_root):
        entry_path = os.path.join(data_root, entry)
        # Legacy flat layout: data/2026-W10/
        if re.match(r"^\d{4}-W\d{2}$", entry) and os.path.isdir(entry_path):
            periods.add(entry)
        # Hierarchical layout: data/2026/W10/
        elif re.match(r"^\d{4}$", entry) and os.path.isdir(entry_path):
            for week in os.listdir(entry_path):
                if re.match(r"^W\d{2}$", week) and os.path.isdir(os.path.join(entry_path, week)):
                    periods.add(f"{entry}-{week}")
    return sorted(periods)


def _generate_period(period: str, base: str) -> None:
    data_dir = fleet_logic.resolve_data_dir(period, base)
    if not os.path.isdir(data_dir):
        print(f"  [skip] {period} - no data found in {data_dir}")
        return

    year, week = period.split("-")
    out_dir = os.path.join(base, "reports", year, week)
    os.makedirs(out_dir, exist_ok=True)

    df      = fleet_logic.load_period(data_dir)
    drivers = fleet_logic.drivers_in(df)

    for d in drivers:
        fname = f"{d.replace(' ', '_')}.xlsx"
        write_driver_report(d, df, period, os.path.join(out_dir, fname))

    write_fleet_summary(df, period, os.path.join(out_dir, "_fleet_summary.xlsx"))
    print(f"  {period}  ->  {len(drivers):2d} drivers  ->  {out_dir}")


def main() -> None:
    base = os.path.dirname(os.path.abspath(__file__))

    if len(sys.argv) > 1 and sys.argv[1] == "--all":
        periods = _find_all_periods(base)
        if not periods:
            print("No data periods found under data/.")
            sys.exit(1)
        print(f"Found {len(periods)} period(s) - generating all...")
        for p in periods:
            _generate_period(p, base)
        print(f"\nDone. Reports written to reports/"  )
    else:
        period = sys.argv[1] if len(sys.argv) > 1 else "2026-W10"
        print(f"Generating {period}...")
        _generate_period(period, base)


if __name__ == "__main__":
    main()
